import openpyxl

wb = openpyxl.load_workbook('d:/projects/porfolio/japaneseStyle/404130_rirekisho_format.xlsx')
ws = wb.active

data = {
    # DATE
    'E3': '2026年 3月 14日現在',
    
    # PERSONAL INFO
    'D5': 'グッデティ ガウラヴ',
    'D7': 'Guddeti Gaurav',
    'G13': '2005年 8月 5日生 （満 20 歳）',
    'L13': '男',
    'M15': '07276082005',
    'D17': '4110048',
    'D18': 'Pune B2-902, Kumar Prithvi Achal Nagar Society',
    'M18': 'guddetigaurav1@gmail.com',
    'D23': '', # Address 2
    
    # EDUCATION / WORK (Left Side)
    'C29': '                    学    歴',
    
    'A31': '2021', 'B31': '6', 'C31': 'Vishwakarma University of Commerce, Arts and Science HSC　入学',
    'A32': '2023', 'B32': '4', 'C32': 'Vishwakarma University of Commerce, Arts and Science HSC　卒業',
    'A33': '2023', 'B33': '8', 'C33': 'Vishwakarma University AI & ML　入学',
    'A34': '2027', 'B34': '4', 'C34': 'Vishwakarma University AI & ML　卒業見込み',
    
    'C36': '                    職    歴',
    'A38': '2025', 'B38': '6', 'C38': 'ChatMaven　入社　AI Dev （アルバイト）',
    'A39': '2025', 'B39': '10', 'C39': '契約期間満了により退職',
    'C41': '    現在に至る',
    'M41': '以    上',
    
    # CERTIFICATIONS (Right Side)
    'O23': '2023', 'P23': '3', 'Q23': 'Generative AI Fundamentals',
    'O24': '2023', 'P24': '5', 'Q24': 'AI For India 2.0',
    'O25': '2023', 'P25': '4', 'Q25': 'The Complete Python Bootcamp',
    'O26': '2025', 'P26': '8', 'Q26': 'IBM Skills Network (Coursera)',
    
    # SELF PR (Right Side)
    'O40': '''私は、デモだけでなく実世界で機能するものを作ります。ChatMaven.aiでのインターンシップでは、RAGベースのAIシステムを本番環境にデプロイし、理論と実践が全く異なるスキルであることを素早く学びました。私は両方を身につけています。
私の仕事はフルスタックに及びます：モバイルアプリ（KhaoozyはKotlinを使用してAndroidで動作）、音声パイプライン（RelayXはSTT/TTSでライブコールを処理）、そしてメモリ対応AIエージェント（Echoは私がゼロから設計したカスタムRAGエンジンを使用）。私は一つの層だけに特化せず、それらがどのように接続されているかを理解しています。これにより、クロスファンクショナルチームで役立つ存在となっています。
デリーの国際会議で発表されたマルチモーダル生体認証セキュリティに関する論文を共著し、これにより自分が構築するシステムについてより厳密に書き、考えるようになりました。この研究マインドセットは現在、プロジェクトへのアプローチに表れており、コーディングに飛びつく前に問題をフレーム化します。
仕事以外では、主要な大学のハッカソンのモデレーターを務め、そのために金メダルを獲得しました。また、全国レベルのValorantの競技者でもあります。これは無関係に聞こえるかもしれませんが、プレッシャーの下でのハイステークスなチームコーディネーションは、実際に転用可能な本物のスキルです。
私はVishwakarma大学で8.15のCGPAで学んでいますが、実際の学びは構築を通じて行っています。私のGitHubは、成績証明書よりも私について多くを語っています。最新のプロジェクトをご覧ください。'''
}

for coord, value in data.items():
    try:
        cell = ws[coord]
        # Check if cell is a MergedCell (read-only)
        if isinstance(cell, openpyxl.cell.cell.MergedCell):
            # Find the top-left cell of the merged range this cell belongs to
            for merged_range in ws.merged_cells.ranges:
                if coord in merged_range:
                    top_left = merged_range.start_cell.coordinate
                    ws[top_left].value = value
                    break
        else:
            cell.value = value
    except Exception as e:
        print(f"Error setting {coord}: {e}")

wb.save('d:/projects/porfolio/japaneseStyle/gaurav_rirekisho.xlsx')
print("Excel file successfully written to gaurav_rirekisho.xlsx")
